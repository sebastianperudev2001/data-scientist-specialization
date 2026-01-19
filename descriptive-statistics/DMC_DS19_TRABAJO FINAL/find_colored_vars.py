
import openpyxl
import pandas as pd

def find_colored_variables(file_path, sheet_name):
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        if sheet_name not in wb.sheetnames:
            print(f"Sheet '{sheet_name}' not found.")
            return

        sheet = wb[sheet_name]
        colored_vars = []
        
        # Assuming header is in the first row, data starts from second
        # We look at the first column (A) for variable names, or checking all columns?
        # Usually "variables" are in a specific column. Let's look at column A or B based on previous notebook view.
        # Notebook output showed: VARIABLE, DESCRIPCIÓN, TIPO DE VARIABLE
        # VARIABLE is likely column B (index 2) or A? 
        # Notebook output: index 0 is AÑO, 1 is P400N...
        # Let's inspect the header row to find "VARIABLE" column index.
        
        variable_col_idx = None
        header_row = 1
        
        for cell in sheet[header_row]:
            if cell.value and "VARIABLE" in str(cell.value).upper():
                variable_col_idx = cell.column
                break
        
        if variable_col_idx is None:
            # Fallback: assume column B (2) if A is index? Or A (1).
            # The pandas output showed an index (0, 1...) then VARIABLE column.
            # In Excel, it's likely the first or second column.
            # Let's check the first few non-empty cells in row 1.
            variable_col_idx = 2 # Guessing B based on common formats, but let's be broader.
            print("Could not find 'VARIABLE' header, checking all cells for colors...")

        print(f"Scanning for colored cells in sheet '{sheet_name}'...")
        
        # Iterate through rows
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                # Check for fill color
                if cell.fill and cell.fill.start_color.index:
                    # '00000000' is usually transparent/no color in openpyxl if type is 'none' but sometimes it's index 00.
                    # Standard check:
                    color = cell.fill.start_color.index
                    # 64 is often default/auto. 
                    if color != '00000000' and color != 64: 
                         # We found a colored cell.
                         # Try to associate it with a variable name in that row.
                         # If we found the variable column, get that value.
                         # Else, just print the cell value.
                         
                         var_name = "Unknown"
                         if variable_col_idx:
                             # Get value from the variable column in this row
                             var_cell = sheet.cell(row=cell.row, column=variable_col_idx)
                             var_name = var_cell.value
                         else:
                             var_name = cell.value
                             
                         if var_name:
                             colored_vars.append({
                                 "row": cell.row,
                                 "col": cell.column,
                                 "value": cell.value,
                                 "variable_associated": var_name,
                                 "color": color
                             })

        # Deduplicate by variable name
        unique_vars = {}
        for item in colored_vars:
            v_name = item['variable_associated']
            if v_name not in unique_vars:
                unique_vars[v_name] = item

        print(f"Found {len(unique_vars)} unique colored variables/rows.")
        for v in unique_vars.values():
            print(f"Variable: {v['variable_associated']} (Value in colored cell: {v['value']})")

    except Exception as e:
        print(f"Error: {e}")

if __name__ == "__main__":
    find_colored_variables('data.xlsx', 'Diccionario Enaho 400')
